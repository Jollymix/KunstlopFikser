import os
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText
from datetime import datetime, timedelta
import subprocess
import tempfile
from io import BytesIO
import random
import re
import json


def decode_xml_bytes(data):
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError:
        return data.decode("cp1252", errors="replace")


def safe_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def is_cancelled(status):
    return "avmeld" in (status or "").strip().lower()


def parse_competition(xml_text, log):
    rows = []
    roots = []
    try:
        roots = [ET.fromstring(xml_text)]
    except ET.ParseError as exc:
        if "junk after document element" in str(exc):
            log("XML inneholder flere dokumenter, prøver å hente alle OdfBody.")
            for match in re.finditer(r"<OdfBody\\b", xml_text):
                start = match.start()
                end = xml_text.find("</OdfBody>", start)
                if end != -1:
                    chunk = xml_text[start : end + len("</OdfBody>")]
                    try:
                        roots.append(ET.fromstring(chunk))
                    except ET.ParseError:
                        continue
        else:
            raise

    for root in roots:
        comp = root.find("Competition")
        if comp is None:
            continue

        for participant in comp.findall("Participant"):
            given = (participant.attrib.get("GivenName") or "").strip()
            family = (participant.attrib.get("FamilyName") or "").strip()
            print_name = (participant.attrib.get("PrintName") or "").strip()
            gender = (participant.attrib.get("Gender") or "").strip()
            org = (participant.attrib.get("Organisation") or "").strip()
            participant_code = (participant.attrib.get("Code") or "").strip()

            for discipline in participant.findall("Discipline"):
                for reg in discipline.findall("RegisteredEvent"):
                    event_code = (reg.attrib.get("Event") or "").strip()
                    entry_order = ""
                    music = {}
                    clubs = {}
                    elements_free = []
                    elements_short = []

                    for entry in reg.findall("EventEntry"):
                        code = (entry.attrib.get("Code") or "").strip()
                        pos = safe_int(entry.attrib.get("Pos"))
                        val = (entry.attrib.get("Value") or "").strip()

                        if code == "ENTRY_ORDER":
                            entry_order = val
                        elif code == "MUSIC":
                            if pos:
                                music[pos] = val
                        elif code == "CLUB":
                            if pos:
                                clubs[pos] = val
                        elif code == "ELEMENT_CODE_FREE":
                            elements_free.append((pos, val))
                        elif code == "ELEMENT_CODE_SHORT":
                            elements_short.append((pos, val))

                    elements_free = [v for _, v in sorted(elements_free) if v]
                    elements_short = [v for _, v in sorted(elements_short) if v]

                    log(f"Leser deltager: {print_name} (Event: {event_code})")
                    rows.append(
                        {
                            "PrintName": print_name,
                            "GivenName": given,
                            "FamilyName": family,
                            "Gender": gender,
                            "Organisation": org,
                            "ParticipantCode": participant_code,
                            "Event": event_code,
                            "EntryOrder": entry_order,
                            "Music1": music.get(1, ""),
                            "Music2": music.get(2, ""),
                            "Club1": clubs.get(1, ""),
                            "Club2": clubs.get(2, ""),
                            "ElementsFree": ", ".join(elements_free),
                            "ElementsShort": ", ".join(elements_short),
                        }
                    )

    return rows


def normalize_name(value):
    return " ".join((value or "").strip().lower().split())


def normalize_text(value):
    text = (value or "").strip().lower()
    text = (
        text.replace("ø", "o")
        .replace("å", "a")
        .replace("æ", "ae")
        .replace("ö", "o")
        .replace("ä", "a")
        .replace("é", "e")
    )
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def tokenize_name(value):
    norm = normalize_text(value)
    return [t for t in norm.split() if t]


def sanitize_filename(value):
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", (value or "").strip())
    return cleaned.strip("._") or "fil"


def get_version():
    version = "ukjent"
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).resolve().parent,
        )
        if result.returncode == 0:
            version = result.stdout.strip() or "ukjent"
    except Exception:
        version = "ukjent"
    return version


def format_generated_ts():
    computer_name = os.environ.get("COMPUTERNAME") or os.environ.get("HOSTNAME") or ""
    suffix = f" â€¢ {computer_name}" if computer_name else ""
    version = get_version()
    return (
        datetime.now().strftime("Generert %d.%m.%Y %H:%M")
        + suffix
        + f" â€¢ Revisjon: {version}"
    )


def name_matches_filename(given, family, filename):
    family_tokens = tokenize_name(family)
    hay = normalize_text(filename)
    if not family_tokens:
        return False
    if all(token in hay for token in family_tokens):
        return True
    return family_tokens[0] in hay


def format_duration(seconds):
    if seconds is None:
        return ""
    total = int(round(seconds))
    mins = total // 60
    secs = total % 60
    return f"{mins}:{secs:02d}"


def parse_time_hhmm(value):
    if not value:
        return None
    parts = [p.strip() for p in value.split(":")]
    try:
        if len(parts) == 2:
            h, m = int(parts[0]), int(parts[1])
            return datetime(2000, 1, 1, h, m, 0)
        if len(parts) == 3:
            h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
            return datetime(2000, 1, 1, h, m, s)
    except ValueError:
        return None
    return None


def parse_duration_mmss(value):
    if not value:
        return None
    parts = [p.strip() for p in value.split(":")]
    try:
        if len(parts) == 2:
            m, s = int(parts[0]), int(parts[1])
            return m * 60 + s
        if len(parts) == 3:
            h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
            return h * 3600 + m * 60 + s
    except ValueError:
        return None
    return None


def parse_date_ddmmyy(value):
    try:
        return datetime.strptime(value.strip(), "%d.%m.%y").date()
    except Exception:
        return None


def format_date_long(date_obj):
    months = [
        "januar",
        "februar",
        "mars",
        "april",
        "mai",
        "juni",
        "juli",
        "august",
        "september",
        "oktober",
        "november",
        "desember",
    ]
    return f"{date_obj.day} {months[date_obj.month - 1]} {date_obj.year}"


def is_registered(status):
    if is_cancelled(status):
        return False
    if not status:
        return False
    text = str(status).strip().lower()
    if "ikke sjekket inn" in text:
        return True
    return "påmeld" in text or "registr" in text or "bekreftet" in text


def load_participants_from_excel(excel_path, log):
    try:
        import openpyxl
    except Exception:
        log("Mangler openpyxl. Installer med: pip install openpyxl")
        return []

    if not Path(excel_path).exists():
        log(f"Fant ikke excel: {excel_path}")
        return []

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    except Exception as exc:
        log(f"Kunne ikke lese excel: {excel_path} ({exc})")
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        log("Excel er tom.")
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    header_map = {h: idx for idx, h in enumerate(headers)}

    def idx(*names):
        for name in names:
            if name in header_map:
                return header_map[name]
        return None

    i_given = idx("Fornavn")
    i_family = idx("Etternavn")
    i_gender = idx("Kjønn")
    i_club = idx("Klubb")
    i_status = idx("Påmelding", "Påmeldingsstatus", "Deltakerstatus")

    if i_given is None or i_family is None:
        log("Finner ikke nødvendige kolonner i excel (Fornavn/Etternavn).")
        return []

    out = []
    for row in rows[1:]:
        given = row[i_given] if i_given < len(row) else ""
        family = row[i_family] if i_family < len(row) else ""
        gender = row[i_gender] if i_gender is not None and i_gender < len(row) else ""
        club = row[i_club] if i_club is not None and i_club < len(row) else ""
        status = row[i_status] if i_status is not None and i_status < len(row) else ""

        if not (given or family):
            continue

        print_name = f"{str(family).strip()}, {str(given).strip()}".strip(", ")
        out.append(
            {
                "PrintName": print_name,
                "NavnFraIsonen": f"{str(given).strip()} {str(family).strip()}".strip(),
                "NavnFraFsm": "",
                "GivenName": (str(given).strip() if given is not None else ""),
                "FamilyName": (str(family).strip() if family is not None else ""),
                "Gender": (str(gender).strip() if gender is not None else ""),
                "Organisation": (str(club).strip() if club is not None else ""),
                "ParticipantCode": "",
                "Event": "",
                "EntryOrder": "",
                "Påmelding": (str(status).strip() if status is not None else ""),
                "Music1": "",
                "Music2": "",
                "Club1": "",
                "Club2": "",
                "ElementsFree": "",
                "ElementsShort": "",
                "Manglende i zip": "",
            }
        )

    log(f"Lest excel: {excel_path} ({len(out)} rader)")
    return out


def parse_officials(xml_text, log):
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return 0
    comp = root.find("Competition")
    if comp is None:
        log("Ingen officials funnet i judges-filen.")
        return 0
    officials = comp.findall("Official")
    log(f"Fant {len(officials)} officials i judges-filen.")
    return len(officials)


def generate_excel(rows, out_path, log):
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
    except Exception:
        log("Mangler openpyxl. Installer med: pip install openpyxl")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participants"

    headers = [
        "PrintName",
        "Organisation",
        "ParticipantCode",
        "Event",
        "Påmelding",
        "Musikk",
        "MusikkTid",
        "Club1",
        "Club2",
        "ElementsFree",
        "ElementsShort",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
        if is_cancelled(row.get("Påmelding", "")):
            from openpyxl.styles import PatternFill

            fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.font = Font(color="7A0B0B")

    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 18

    try:
        wb.save(out_path)
    except PermissionError:
        log(f"Kunne ikke skrive Excel (filen er trolig åpen): {out_path}")
        return False
    except Exception as exc:
        log(f"Kunne ikke skrive Excel: {out_path} ({exc})")
        return False
    log(f"Excel skrevet: {out_path}")
    return True


def generate_html(rows, out_path, title, log):
    headers = [
        "PrintName",
        "Organisation",
        "ParticipantCode",
        "Event",
        "Påmelding",
        "Musikk",
        "MusikkTid",
        "Club1",
        "Club2",
        "ElementsFree",
        "ElementsShort",
    ]

    def esc(s):
        return (
            str(s)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    rows_html = []
    for row in rows:
        row_style = ""
        if is_cancelled(row.get("Påmelding", "")):
            row_style = ' style="background:#f8d7da;color:#7a0b0b;"'
        cells = "".join(f"<td>{esc(row.get(h, ''))}</td>" for h in headers)
        rows_html.append(f"<tr{row_style}>{cells}</tr>")

    html = f"""<!doctype html>
<html lang="no">
<head>
  <meta charset="utf-8">
  <title>{esc(title)}</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 24px; }}
    table {{ border-collapse: collapse; width: 100%; }}
    th, td {{ border: 1px solid #ccc; padding: 6px 8px; text-align: left; }}
    th {{ background: #f4f4f4; }}
  </style>
</head>
<body>
  <h2>{esc(title)}</h2>
  <table>
    <thead>
      <tr>
        {''.join(f'<th>{esc(h)}</th>' for h in headers)}
      </tr>
    </thead>
    <tbody>
      {''.join(rows_html)}
    </tbody>
  </table>
</body>
</html>
"""

    try:
        Path(out_path).write_text(html, encoding="utf-8")
    except PermissionError:
        log(f"Kunne ikke skrive HTML (filen er trolig åpen): {out_path}")
        return False
    except Exception as exc:
        log(f"Kunne ikke skrive HTML: {out_path} ({exc})")
        return False
    log(f"HTML skrevet: {out_path}")
    return True


def generate_pdf(rows, out_path, title, log):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
            Image,
        )
    except Exception:
        log("Mangler reportlab. Installer med: pip install reportlab")
        return False

    headers = [
        "PrintName",
        "Organisation",
        "Event",
        "Påmelding",
        "Musikk",
        "MusikkTid",
        "ElementsFree",
    ]
    data = [headers]
    for row in rows:
        data.append([row.get(h, "") for h in headers])

    doc = SimpleDocTemplate(out_path, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Loddefjord IL Kunstløp", styles["Title"]),
        Spacer(1, 6),
    ]
    logo_path = Path(__file__).resolve().parent / "Lil Logo.jpg"
    if logo_path.exists():
        try:
            logo = Image(str(logo_path))
            logo.drawHeight = 50
            logo.drawWidth = 50 * (logo.imageWidth / logo.imageHeight)
            story.append(logo)
            story.append(Spacer(1, 6))
        except Exception:
            log("Klarte ikke å lese logo-bildet.")
    story.append(Paragraph(title, styles["Heading2"]))
    story.append(Spacer(1, 12))

    table = Table(data, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]
    for idx, row in enumerate(rows, start=1):
        if is_cancelled(row.get("Påmelding", "")):
            style_cmds.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#F8D7DA")))
            style_cmds.append(("TEXTCOLOR", (0, idx), (-1, idx), colors.HexColor("#7A0B0B")))
    table.setStyle(TableStyle(style_cmds))
    story.append(table)
    story.append(Spacer(1, 6))
    story.append(Paragraph(format_generated_ts(), styles["Normal"]))
    try:
        doc.build(story)
    except PermissionError:
        log(f"Kunne ikke skrive PDF (filen er trolig åpen): {out_path}")
        return False
    except Exception as exc:
        log(f"Kunne ikke skrive PDF: {out_path} ({exc})")
        return False
    log(f"PDF skrevet: {out_path}")
    return True


def build_startliste(
    rows,
    group_size,
    interval_seconds,
    start_time,
    warmup_seconds=0,
    pause_after=None,
    pause_seconds=None,
    pause_label="Vanningspause",
):
    entries = []
    if not rows:
        return entries
    group_size = max(1, group_size)
    interval_seconds = max(1, interval_seconds)
    pause_after = pause_after if pause_after and pause_after > 0 else None
    pause_seconds = pause_seconds if pause_seconds and pause_seconds > 0 else None

    index = 0
    group_num = 1
    current_dt = start_time
    while index < len(rows):
        group_start_dt = current_dt
        group_label_time = group_start_dt.strftime("%H:%M:%S")
        if group_num > 1:
            group_label_time = f"ca. {group_label_time}"
        warmup_end = group_start_dt + timedelta(seconds=warmup_seconds)
        group_entry = {
            "is_group": True,
            "start": group_label_time,
            "end": warmup_end.strftime("%H:%M:%S"),
            "nr": "",
            "navn": f"Oppvarmingsgruppe {group_num}",
            "klubb": "",
        }
        entries.append(group_entry)

        current_dt = warmup_end
        group_rows = rows[index : index + group_size]
        for offset, row in enumerate(group_rows, start=1):
            runner_start = current_dt
            runner_end = runner_start + timedelta(seconds=interval_seconds)
            entries.append(
                {
                    "is_group": False,
                    "start": runner_start.strftime("%H:%M:%S"),
                    "end": runner_end.strftime("%H:%M:%S"),
                    "nr": index + offset,
                    "navn": f"{row.get('GivenName', '')} {row.get('FamilyName', '')}".strip(),
                    "klubb": row.get("Organisation", ""),
                }
            )
            current_dt = runner_end
            if pause_after and pause_seconds and (index + offset) == pause_after:
                pause_start = current_dt
                pause_end = pause_start + timedelta(seconds=pause_seconds)
                entries.append(
                    {
                        "is_group": True,
                        "start": pause_start.strftime("%H:%M:%S"),
                        "end": pause_end.strftime("%H:%M:%S"),
                        "nr": "",
                        "navn": pause_label,
                        "klubb": "",
                    }
                )
                current_dt = pause_end

        index += group_size
        group_num += 1
    return entries


def generate_startliste_excel(entries, out_path, title, log):
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill
    except Exception:
        log("Mangler openpyxl. Installer med: pip install openpyxl")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Startliste"

    ws.append([title])
    ws.append([])
    headers = ["Nr.", "Start", "", "Slutt", "Navn", "Klubb"]
    ws.append(headers)

    header_font = Font(bold=True)
    for cell in ws[3]:
        cell.font = header_font

    header_fill = PatternFill(start_color="ADADAD", end_color="ADADAD", fill_type="solid")
    group_fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
    group_font = Font(bold=True)

    for entry in entries:
        ws.append([entry["nr"], entry["start"], "-", entry["end"], entry["navn"], entry["klubb"]])
        if entry["is_group"]:
            for cell in ws[ws.max_row]:
                cell.font = group_font
                cell.fill = group_fill
    for cell in ws[3]:
        cell.fill = header_fill

    ws.column_dimensions[get_column_letter(1)].width = 6
    ws.column_dimensions[get_column_letter(2)].width = 10
    ws.column_dimensions[get_column_letter(3)].width = 3
    ws.column_dimensions[get_column_letter(4)].width = 10
    ws.column_dimensions[get_column_letter(5)].width = 38
    ws.column_dimensions[get_column_letter(6)].width = 16

    generated_ts = format_generated_ts()
    ws.append([])
    ws.append([generated_ts])

    try:
        wb.save(out_path)
    except PermissionError:
        log(f"Kunne ikke skrive Excel (filen er trolig åpen): {out_path}")
        return False
    except Exception as exc:
        log(f"Kunne ikke skrive Excel: {out_path} ({exc})")
        return False

    log(f"Startliste Excel skrevet: {out_path}")
    return True


def generate_startliste_pdf(entries, out_path, title, log):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception:
        log("Mangler reportlab. Installer med: pip install reportlab")
        return False

    font_name = "Helvetica"
    try:
        calibri_path = Path("C:/Windows/Fonts/calibri.ttf")
        if calibri_path.exists():
            pdfmetrics.registerFont(TTFont("Calibri", str(calibri_path)))
            font_name = "Calibri"
    except Exception:
        font_name = "Helvetica"

    data = [["Nr.", "Start", "", "Slutt", "Navn", "Klubb"]]
    body_style = ParagraphStyle(
        "BodyCell",
        fontName=font_name,
        fontSize=10,
        leading=11,
    )
    for entry in entries:
        name_cell = Paragraph(entry["navn"], body_style)
        club_cell = Paragraph(entry["klubb"], body_style)
        data.append([entry["nr"], entry["start"], "-", entry["end"], name_cell, club_cell])

    doc = SimpleDocTemplate(out_path, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "StartTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=16,
        leading=18,
    )
    generated_ts = format_generated_ts()
    story = [Paragraph(title, title_style), Spacer(1, 8)]

    col_widths = [28, 64, 10, 64, 230, 127]
    table = Table(data, repeatRows=1, colWidths=col_widths)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ADADAD")),
        ("FONTNAME", (0, 0), (-1, 0), font_name),
        ("FONTNAME", (0, 1), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("ALIGN", (0, 0), (3, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.black),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.black),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.black),
        ("LINEBEFORE", (0, 0), (0, -1), 0.5, colors.black),
        ("LINEAFTER", (-1, 0), (-1, -1), 0.5, colors.black),
    ]
    row_idx = 1
    for entry in entries:
        if entry["is_group"]:
            style_cmds.append(
                ("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#D0D0D0"))
            )
            style_cmds.append(("FONTNAME", (0, row_idx), (-1, row_idx), "Helvetica-Bold"))
        row_idx += 1
    table.setStyle(TableStyle(style_cmds))
    story.append(table)
    story.append(Spacer(1, 6))
    story.append(Paragraph(generated_ts, ParagraphStyle("Gen", fontName=font_name, fontSize=9)))

    try:
        doc.build(story)
    except PermissionError:
        log(f"Kunne ikke skrive PDF (filen er trolig åpen): {out_path}")
        return False
    except Exception as exc:
        log(f"Kunne ikke skrive PDF: {out_path} ({exc})")
        return False

    log(f"Startliste PDF skrevet: {out_path}")
    return True


def generate_vlc_playlist(rows, out_dir, base_name, music_zip, log):
    if not music_zip:
        log("Ingen musikk-zip funnet, kan ikke lage spilleliste.")
        return False

    playlist_path = Path(out_dir) / f"Startliste_{base_name}.m3u"
    music_out = Path(out_dir) / "music"
    music_out.mkdir(parents=True, exist_ok=True)

    try:
        with zipfile.ZipFile(music_zip, "r") as mz:
            lines = ["#EXTM3U"]
            for row in rows:
                fname = row.get("MusikkFil", "")
                if not fname:
                    continue
                dest_path = music_out / fname
                dest_path.parent.mkdir(parents=True, exist_ok=True)
                if not dest_path.exists():
                    try:
                        mz.extract(fname, music_out)
                    except Exception:
                        continue
                duration = row.get("MusikkSek", "")
                performer = row.get("PrintName", "")
                song = Path(fname).stem
                title = f"{performer} - {song}".strip(" -")
                extinf = duration if duration != "" else -1
                lines.append(f"#EXTINF:{extinf},{title}")
                lines.append(str(dest_path.resolve()))
    except Exception as exc:
        log(f"Kunne ikke lage spilleliste: {exc}")
        return False

    playlist_path.write_text("\n".join(lines), encoding="utf-8")
    log(f"Spilleliste skrevet: {playlist_path}")
    return True


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("FSM Data Dekoder")
        self.root.minsize(900, 620)
        self.menubar = tk.Menu(self.root)
        self.root.config(menu=self.menubar)
        self.menu_startliste = tk.Menu(self.menubar, tearoff=0)
        self.menu_rekkefolge = tk.Menu(self.menubar, tearoff=0)
        self.menu_hjelp = tk.Menu(self.menubar, tearoff=0)
        self.menubar.add_cascade(label="Startliste", menu=self.menu_startliste)
        self.menubar.add_cascade(label="Rekkefølge", menu=self.menu_rekkefolge)
        self.menubar.add_cascade(label="Hjelp", menu=self.menu_hjelp)
        self.menu_startliste.add_command(label="Startliste...", command=self.open_startliste_window, state="disabled")
        self.menu_rekkefolge.add_command(label="Lagre rekkefølge", command=self.save_order, state="disabled")
        self.menu_rekkefolge.add_command(label="Last rekkefølge", command=self.load_order, state="disabled")
        self.menu_hjelp.add_command(label="Om", command=self.show_about)
        style = ttk.Style()
        try:
            style.theme_use("vista")
        except tk.TclError:
            style.theme_use("clam")
        style.configure("Title.TLabel", font=("Segoe UI", 14, "bold"))
        style.configure("Subtitle.TLabel", font=("Segoe UI", 9))
        style.configure("Clock.TLabel", font=("Consolas", 56, "bold"))
        self.rows = []
        self.zip_path = None
        self.music_zip = None
        self.music_cache_dir = None
        self.audio_backend = None
        self.audio_ready = False
        self.current_track = ""
        self.is_paused = False
        self.current_duration = 0
        self.play_pause_text = tk.StringVar(value="Spill")
        self.external_playback = False
        self.startliste_window = None

        base_dir = Path(__file__).resolve().parent
        self.folder_var = tk.StringVar(value=str(base_dir))

        container = ttk.Frame(root, padding=10)
        container.pack(fill="both", expand=True)

        header = ttk.Frame(container)
        header.pack(fill="x", pady=(0, 8))
        header_left = ttk.Frame(header)
        header_left.pack(side="left")
        # title removed
        self.clock_var = tk.StringVar(value="00:00:00")
        self.clock_label = ttk.Label(
            header, textvariable=self.clock_var, style="Clock.TLabel", anchor="center"
        )
        self.clock_label.pack(side="left", fill="x", expand=True)

        folder_frame = ttk.Labelframe(container, text="Kilde")
        folder_frame.pack(fill="x", pady=6)
        ttk.Label(folder_frame, text="Mappe med zip:").pack(side="left", padx=(8, 4))
        ttk.Entry(folder_frame, textvariable=self.folder_var, width=60).pack(
            side="left", padx=6
        )
        ttk.Button(folder_frame, text="Velg mappe", command=self.choose_folder).pack(
            side="left"
        )
        ttk.Button(folder_frame, text="Skann filer", command=self.read_zip).pack(
            side="left", padx=(8, 0)
        )
        self.ind_excel = tk.Label(
            folder_frame,
            text="isonen/deltakerliste",
            bg="#cccccc",
            fg="#000000",
            padx=6,
            pady=2,
        )
        self.ind_music = tk.Label(
            folder_frame,
            text="musikkzip",
            bg="#cccccc",
            fg="#000000",
            padx=6,
            pady=2,
        )
        self.ind_data = tk.Label(
            folder_frame,
            text="FMS data",
            bg="#cccccc",
            fg="#000000",
            padx=6,
            pady=2,
        )
        self.count_label = ttk.Label(folder_frame, text="Utøvere: 0")
        self.count_label.pack(side="right", padx=8)
        self.ind_excel.pack(side="right", padx=4)
        self.ind_music.pack(side="right", padx=4)
        self.ind_data.pack(side="right", padx=4)

        btn_frame = ttk.Frame(container)
        btn_frame.pack(fill="x", pady=4)

        table_frame = ttk.Labelframe(container, text="Utøvere")
        table_frame.pack(fill="both", expand=True, pady=6)
        columns = (
            "startnummer",
            "starttid",
            "navn_isonen",
            "navn_fsm",
            "klubb",
            "påmelding",
            "musikknavn",
            "musikktid",
        )
        self.tree = ttk.Treeview(
            table_frame, columns=columns, show="headings", height=10
        )
        self.tree.heading("startnummer", text="Startnummer")
        self.tree.heading("starttid", text="Starttid")
        self.tree.heading("navn_isonen", text="Navn fra isonen")
        self.tree.heading("navn_fsm", text="Navn fra fsm")
        self.tree.heading("klubb", text="Klubb")
        self.tree.heading("påmelding", text="Påmelding")
        self.tree.heading("musikknavn", text="MP3-fil")
        self.tree.heading("musikktid", text="MusikkTid")
        self.tree.column("startnummer", width=90, anchor="center")
        self.tree.column("starttid", width=90, anchor="center")
        self.tree.column("navn_isonen", width=200, anchor="w")
        self.tree.column("navn_fsm", width=200, anchor="w")
        self.tree.column("klubb", width=140, anchor="w")
        self.tree.column("påmelding", width=120, anchor="w")
        self.tree.column("musikknavn", width=260, anchor="w")
        self.tree.column("musikktid", width=80, anchor="center")
        self.tree.configure(selectmode="browse")
        self.tree.tag_configure("missing_music", foreground="#b00020")
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        controls_frame = ttk.Frame(table_frame)
        self.btn_move_up = ttk.Button(
            controls_frame, text="Flytt opp", command=self.move_selected_up, state="disabled"
        )
        self.btn_move_down = ttk.Button(
            controls_frame, text="Flytt ned", command=self.move_selected_down, state="disabled"
        )
        self.btn_shuffle = ttk.Button(
            controls_frame, text="Randomiser", command=self.shuffle_rows, state="disabled"
        )
        self.btn_sort_given = ttk.Button(
            controls_frame, text="Sorter fornavn", command=self.sort_by_given, state="disabled"
        )
        self.btn_sort_family = ttk.Button(
            controls_frame, text="Sorter etternavn", command=self.sort_by_family, state="disabled"
        )
        self.btn_move_up.pack(fill="x", padx=6, pady=(6, 2))
        self.btn_move_down.pack(fill="x", padx=6, pady=2)
        self.btn_shuffle.pack(fill="x", padx=6, pady=2)
        self.btn_sort_given.pack(fill="x", padx=6, pady=2)
        self.btn_sort_family.pack(fill="x", padx=6, pady=2)
        # rekkefølge flyttet til meny
        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)
        self.tree.pack(side="left", fill="both", expand=True, padx=6, pady=6)
        controls_frame.pack(side="right", fill="y", pady=6)
        yscroll.pack(side="right", fill="y", pady=6)

        player_frame = ttk.Labelframe(container, text="Avspiller")
        player_frame.pack(fill="x", pady=6)
        player_left = ttk.Frame(player_frame)
        player_left.pack(side="left", padx=6, pady=6)
        self.btn_player_play_pause = ttk.Button(
            player_left,
            textvariable=self.play_pause_text,
            command=self.toggle_play_pause,
            state="disabled",
            width=10,
        )
        self.btn_player_stop = ttk.Button(
            player_left,
            text="Stopp",
            command=self.stop_playback,
            state="disabled",
            width=8,
        )
        self.btn_player_play_pause.pack(side="left", padx=(0, 6))
        self.btn_player_stop.pack(side="left")

        player_mid = ttk.Frame(player_frame)
        player_mid.pack(side="left", fill="x", expand=True, padx=6, pady=6)
        self.player_track_var = tk.StringVar(value="Ingen spor valgt")
        self.player_time_var = tk.StringVar(value="0:00 / 0:00")
        ttk.Label(player_mid, textvariable=self.player_track_var).pack(anchor="w")
        self.player_progress = ttk.Progressbar(player_mid, mode="determinate")
        self.player_progress.pack(fill="x", pady=4)
        ttk.Label(player_mid, textvariable=self.player_time_var).pack(anchor="w")

        log_frame = ttk.Labelframe(container, text="Logg")
        log_frame.pack(fill="x", pady=6)
        self.log_widget = ScrolledText(log_frame, height=6, wrap="word")
        self.log_widget.pack(fill="both", expand=True, padx=6, pady=6)

        today_str = datetime.now().strftime("%d.%m.%y")
        self.start_date_var = tk.StringVar(value=today_str)
        self.start_time_var = tk.StringVar(value="18:00")
        self.interval_var = tk.StringVar(value="3:40")
        self.group_size_var = tk.StringVar(value="8")
        self.location_var = tk.StringVar(value="Iskanten")
        self.warmup_var = tk.StringVar(value="4:00")
        self.pause_after_var = tk.StringVar(value="")
        self.pause_duration_var = tk.StringVar(value="")
        self.pause_label_var = tk.StringVar(value="Vanningspause")
        self.playlist_var = tk.BooleanVar(value=False)

        out_frame = ttk.Labelframe(container, text="Rapporter")
        out_frame.pack(fill="x", pady=6)
        self.var_pdf = tk.BooleanVar(value=False)
        self.var_excel = tk.BooleanVar(value=True)
        self.var_html = tk.BooleanVar(value=True)
        self.chk_pdf = ttk.Checkbutton(out_frame, text="PDF", variable=self.var_pdf)
        self.chk_excel = ttk.Checkbutton(out_frame, text="Excel", variable=self.var_excel)
        self.chk_html = ttk.Checkbutton(out_frame, text="HTML", variable=self.var_html)
        self.chk_pdf.pack(side="left", padx=6, pady=6)
        self.chk_excel.pack(side="left", padx=6, pady=6)
        self.chk_html.pack(side="left", padx=6, pady=6)
        self.btn_generate = ttk.Button(
            out_frame, text="Lag filer", command=self.generate_files, state="disabled"
        )
        self.btn_generate.pack(side="right", padx=6, pady=6)

        self.set_output_controls(enabled=False)
        self.set_table_controls(enabled=False)
        self.update_player_ui()
        self.update_clock()

    def log(self, msg):
        self.log_widget.insert("end", msg + "\n")
        self.log_widget.see("end")
        self.root.update_idletasks()

    def set_output_controls(self, enabled):
        state = "normal" if enabled else "disabled"
        self.chk_pdf.config(state=state)
        self.chk_excel.config(state=state)
        self.chk_html.config(state=state)
        self.btn_generate.config(state=state)
        self.menu_startliste.entryconfig(0, state=state)
    
    def set_table_controls(self, enabled):
        state = "normal" if enabled else "disabled"
        self.btn_move_up.config(state=state)
        self.btn_move_down.config(state=state)
        self.btn_shuffle.config(state=state)
        self.btn_sort_given.config(state=state)
        self.btn_sort_family.config(state=state)
        self.btn_player_play_pause.config(state=state)
        self.btn_player_stop.config(state=state)
        self.menu_rekkefolge.entryconfig(0, state=state)
        self.menu_rekkefolge.entryconfig(1, state=state)

    def open_startliste_window(self):
        if self.startliste_window and self.startliste_window.winfo_exists():
            self.startliste_window.deiconify()
            self.startliste_window.lift()
            return

        win = tk.Toplevel(self.root)
        win.title("Startliste")
        win.resizable(False, False)
        self.startliste_window = win

        frame = ttk.Frame(win, padding=12)
        frame.pack(fill="both", expand=True)

        row = 0
        ttk.Label(frame, text="Dato:").grid(row=row, column=0, sticky="w", padx=(0, 4), pady=4)
        try:
            from tkcalendar import DateEntry

            self.date_widget = DateEntry(
                frame,
                textvariable=self.start_date_var,
                date_pattern="dd.mm.yy",
                width=10,
            )
            self.date_widget.grid(row=row, column=1, sticky="w", pady=4)
        except Exception:
            self.date_widget = None
            ttk.Entry(frame, textvariable=self.start_date_var, width=10).grid(
                row=row, column=1, sticky="w", pady=4
            )

        ttk.Label(frame, text="Start kl:").grid(row=row, column=2, sticky="w", padx=(12, 4))
        ttk.Entry(frame, textvariable=self.start_time_var, width=8).grid(
            row=row, column=3, sticky="w"
        )

        ttk.Label(frame, text="Intervall:").grid(row=row, column=4, sticky="w", padx=(12, 4))
        ttk.Entry(frame, textvariable=self.interval_var, width=6).grid(
            row=row, column=5, sticky="w"
        )

        row += 1
        ttk.Label(frame, text="Gruppe str:").grid(row=row, column=0, sticky="w", padx=(0, 4), pady=4)
        ttk.Entry(frame, textvariable=self.group_size_var, width=4).grid(
            row=row, column=1, sticky="w", pady=4
        )

        ttk.Label(frame, text="Oppvarming:").grid(row=row, column=2, sticky="w", padx=(12, 4))
        ttk.Entry(frame, textvariable=self.warmup_var, width=6).grid(
            row=row, column=3, sticky="w"
        )

        ttk.Label(frame, text="Sted:").grid(row=row, column=4, sticky="w", padx=(12, 4))
        ttk.Entry(frame, textvariable=self.location_var, width=14).grid(
            row=row, column=5, sticky="w"
        )

        row += 1
        ttk.Label(frame, text="Pause etter nr:").grid(row=row, column=0, sticky="w", padx=(0, 4), pady=4)
        ttk.Entry(frame, textvariable=self.pause_after_var, width=4).grid(
            row=row, column=1, sticky="w", pady=4
        )

        ttk.Label(frame, text="Pause varighet:").grid(row=row, column=2, sticky="w", padx=(12, 4))
        ttk.Entry(frame, textvariable=self.pause_duration_var, width=6).grid(
            row=row, column=3, sticky="w"
        )

        ttk.Label(frame, text="Pause tekst:").grid(row=row, column=4, sticky="w", padx=(12, 4))
        ttk.Entry(frame, textvariable=self.pause_label_var, width=12).grid(
            row=row, column=5, sticky="w"
        )

        row += 1
        ttk.Checkbutton(
            frame,
            text="Spilleliste for VLC musikkavspiller",
            variable=self.playlist_var,
        ).grid(row=row, column=0, columnspan=4, sticky="w", pady=(6, 4))

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=row, column=4, columnspan=2, sticky="e", pady=(6, 4))
        ttk.Button(
            btn_frame,
            text="Lag startliste",
            command=self.generate_startliste,
        ).pack(side="right")

    def choose_folder(self):
        path = filedialog.askdirectory(initialdir=self.folder_var.get())
        if path:
            self.folder_var.set(path)

    def read_zip(self):
        self.log_widget.delete("1.0", "end")
        self.rows = []
        self.zip_path = None
        self.ind_data.config(bg="#cccccc")
        self.ind_music.config(bg="#cccccc")
        self.ind_excel.config(bg="#cccccc")
        self.music_cache_dir = None

        folder = Path(self.folder_var.get())
        if not folder.exists():
            messagebox.showerror("Feil", "Mappen finnes ikke.")
            return

        zips = sorted(folder.glob("*.zip"))
        if not zips:
            messagebox.showerror("Feil", "Fant ingen zip-filer i mappen.")
            return

        data_zips = [
            z
            for z in zips
            if z.name.lower().startswith("fmsdata") or z.name.lower().startswith("fsmdata")
        ]
        music_zips = [z for z in zips if z.name.lower().startswith("musikk")]
        excel_files = [p for p in folder.glob("Deltakerliste*.xlsx")]

        if not data_zips:
            messagebox.showerror("Feil", "Fant ingen FMSData*.zip i mappen.")
            return
        if not music_zips:
            messagebox.showerror("Feil", "Fant ingen Musikk*.zip i mappen.")
            return
        if not excel_files:
            messagebox.showerror("Feil", "Fant ingen Deltakerliste*.xlsx i mappen.")
            return

        self.zip_path = data_zips[0]
        if len(data_zips) > 1:
            self.log(f"Fant flere FMSData-zip. Bruker: {self.zip_path.name}")
        self.ind_data.config(bg="#3fbf5f")

        excel_path = excel_files[0]
        if len(excel_files) > 1:
            self.log(f"Fant flere deltakerlister. Bruker: {excel_path.name}")
        self.ind_excel.config(bg="#3fbf5f")
        self.rows = load_participants_from_excel(excel_path, self.log)
        if not self.rows:
            messagebox.showerror("Feil", "Fant ingen deltakere i excel-filen.")
            self.set_output_controls(enabled=False)
            return

        self.log(f"Leser zip: {self.zip_path.name}")
        zip_rows = []
        with zipfile.ZipFile(self.zip_path, "r") as zf:
            xml_entries = [e for e in zf.infolist() if e.filename.lower().endswith(".xml")]
            if not xml_entries:
                messagebox.showerror("Feil", "Fant ingen xml-filer i zip.")
                return

            for entry in xml_entries:
                self.log(f"Leser fil: {entry.filename}")
                data = zf.read(entry)
                xml_text = decode_xml_bytes(data)
                if "judges" in entry.filename.lower():
                    parse_officials(xml_text, self.log)
                else:
                    zip_rows.extend(parse_competition(xml_text, self.log))

        music_zip = music_zips[0]
        if len(music_zips) > 1:
            self.log(f"Fant flere musikk-zip. Bruker: {music_zip.name}")
        self.ind_music.config(bg="#3fbf5f")
        self.music_zip = music_zip

        music_files = []
        music_durations = {}
        if music_zip:
            self.log(f"Leser musikk-zip: {music_zip.name}")
            try:
                with zipfile.ZipFile(music_zip, "r") as mz:
                    for entry in mz.infolist():
                        if entry.filename.lower().endswith(".mp3"):
                            music_files.append(entry.filename)
                            try:
                                from mutagen.mp3 import MP3

                                data = mz.read(entry)
                                audio = MP3(BytesIO(data))
                                music_durations[entry.filename] = audio.info.length
                            except Exception:
                                music_durations[entry.filename] = None
            except Exception as exc:
                self.log(f"Kunne ikke lese musikk-zip: {music_zip} ({exc})")
        else:
            self.log("Fant ingen musikk-zip (navn med 'MUSIKK').")

        zip_map = {}
        for row in zip_rows:
            key = (
                normalize_name(row.get("GivenName")),
                normalize_name(row.get("FamilyName")),
            )
            if key[0] or key[1]:
                zip_map[key] = row

        excel_keys = set()
        used_music_files = set()
        for row in self.rows:
            key = (
                normalize_name(row.get("GivenName")),
                normalize_name(row.get("FamilyName")),
            )
            excel_keys.add(key)
            zip_row = zip_map.get(key)
            if zip_row:
                row["ParticipantCode"] = zip_row.get("ParticipantCode", "")
                row["Event"] = zip_row.get("Event", "")
                row["EntryOrder"] = zip_row.get("EntryOrder", "")
                row["Music1"] = zip_row.get("Music1", "")
                row["Music2"] = zip_row.get("Music2", "")
                row["Club1"] = zip_row.get("Club1", "")
                row["Club2"] = zip_row.get("Club2", "")
                row["ElementsFree"] = zip_row.get("ElementsFree", "")
                row["ElementsShort"] = zip_row.get("ElementsShort", "")
                zip_print = zip_row.get("PrintName") or f"{zip_row.get('GivenName', '')} {zip_row.get('FamilyName', '')}".strip()
                row["NavnFraFsm"] = zip_print
                row["Manglende i zip"] = ""
            else:
                row["Manglende i zip"] = "JA"
                row["NavnFraFsm"] = ""

            if music_files:
                matched = self.match_music_file(row, music_files, used_music_files)
                row["Musikk"] = "ok" if matched else "mangler"
                row["MusikkFil"] = matched or ""
                musikk_sec = music_durations.get(matched) if matched else None
                if matched and musikk_sec is None:
                    row["MusikkTid"] = "Klarer ikke å hente tid"
                    row["MusikkSek"] = ""
                else:
                    row["MusikkTid"] = format_duration(musikk_sec)
                    row["MusikkSek"] = int(round(musikk_sec)) if musikk_sec else ""
            else:
                row["Musikk"] = "mangler"
                row["MusikkFil"] = ""
                row["MusikkTid"] = ""
                row["MusikkSek"] = ""

        for key, row in zip_map.items():
            if key not in excel_keys:
                zip_given = (row.get("GivenName") or "").strip()
                zip_family = (row.get("FamilyName") or "").strip()
                zip_print = row.get("PrintName") or f"{zip_given} {zip_family}".strip()
                self.rows.append(
                    {
                        "PrintName": zip_print,
                        "NavnFraIsonen": "",
                        "NavnFraFsm": zip_print,
                        "GivenName": zip_given,
                        "FamilyName": zip_family,
                        "Gender": row.get("Gender", ""),
                        "Organisation": row.get("Organisation", ""),
                        "ParticipantCode": row.get("ParticipantCode", ""),
                        "Event": row.get("Event", ""),
                        "EntryOrder": row.get("EntryOrder", ""),
                        "Påmelding": "",
                        "Music1": row.get("Music1", ""),
                        "Music2": row.get("Music2", ""),
                        "Club1": row.get("Club1", ""),
                        "Club2": row.get("Club2", ""),
                        "ElementsFree": row.get("ElementsFree", ""),
                        "ElementsShort": row.get("ElementsShort", ""),
                        "Manglende i zip": "",
                        "Musikk": "mangler",
                        "MusikkFil": "",
                        "MusikkTid": "",
                        "MusikkSek": "",
                        "StartTid": "",
                    }
                )

        if music_files:
            self.log("MP3-filer i musikk-zip:")
            for fname in sorted(music_files):
                self.log(f"- {fname}")

        if not self.rows:
            messagebox.showwarning("Info", "Fant ingen deltakere i xml.")
            self.set_output_controls(enabled=False)
            return

        self.log(f"Totalt deltakere: {len(self.rows)}")
        self.count_label.config(text=f"Utøvere: {len(self.rows)}")
        self.set_output_controls(enabled=True)
        self.set_table_controls(enabled=True)
        self.refresh_table()

    def refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        rows_values = []
        for idx, row in enumerate(self.rows, start=1):
            missing = not row.get("MusikkFil")
            mp3_text = "mangler musikk" if missing else row.get("MusikkFil", "")
            tags = ("missing_music",) if missing else ()
            values = (
                idx,
                row.get("StartTid", ""),
                row.get("NavnFraIsonen", ""),
                row.get("NavnFraFsm", ""),
                row.get("Organisation", ""),
                row.get("Påmelding", ""),
                mp3_text,
                row.get("MusikkTid", ""),
            )
            rows_values.append(values)
            self.tree.insert("", "end", values=values, tags=tags)
        self.autosize_columns(rows_values)

    def autosize_columns(self, rows_values):
        try:
            import tkinter.font as tkfont
        except Exception:
            return
        font = tkfont.nametofont("TkDefaultFont")
        columns = list(self.tree["columns"])
        if not columns:
            return
        padding = 16
        max_widths = {}
        for col in columns:
            heading = self.tree.heading(col).get("text", "")
            max_widths[col] = font.measure(str(heading)) + padding
        for row in rows_values:
            for col, value in zip(columns, row):
                width = font.measure(str(value)) + padding
                if width > max_widths.get(col, 0):
                    max_widths[col] = width
        for col in columns:
            self.tree.column(col, width=max_widths.get(col, 80), stretch=False)

    def update_clock(self):
        self.clock_var.set(datetime.now().strftime("%H:%M:%S"))
        self.root.after(500, self.update_clock)

    def sort_by_given(self):
        if not self.rows:
            return
        self.rows.sort(
            key=lambda r: (
                normalize_text(r.get("GivenName")),
                normalize_text(r.get("FamilyName")),
            )
        )
        self.refresh_table()

    def sort_by_family(self):
        if not self.rows:
            return
        self.rows.sort(
            key=lambda r: (
                normalize_text(r.get("FamilyName")),
                normalize_text(r.get("GivenName")),
            )
        )
        self.refresh_table()

    def on_tree_double_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        column = self.tree.identify_column(event.x)
        if not column:
            return
        col_index = int(column[1:]) - 1
        columns = list(self.tree["columns"])
        if col_index < 0 or col_index >= len(columns):
            return
        if columns[col_index] != "musikknavn":
            return
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return
        values = self.tree.item(item_id, "values")
        if col_index >= len(values):
            return
        filename = values[col_index]
        self.play_mp3_file(filename)

    def get_cached_mp3_path(self, filename):
        if not filename:
            messagebox.showinfo("Info", "Ingen MP3-fil registrert på denne raden.")
            return None
        if not self.music_zip:
            messagebox.showerror("Feil", "Fant ingen musikk-zip.")
            return None
        try:
            with zipfile.ZipFile(self.music_zip, "r") as mz:
                try:
                    data = mz.read(filename)
                except KeyError:
                    messagebox.showerror("Feil", f"Fant ikke MP3 i zip: {filename}")
                    return None
        except Exception as exc:
            messagebox.showerror("Feil", f"Kunne ikke lese musikk-zip: {exc}")
            return None

        if not self.music_cache_dir:
            base = Path(tempfile.gettempdir()) / "fms_gui_music_cache"
            base.mkdir(parents=True, exist_ok=True)
            safe_zip = sanitize_filename(self.music_zip.stem)
            self.music_cache_dir = base / safe_zip
            self.music_cache_dir.mkdir(parents=True, exist_ok=True)

        safe_name = sanitize_filename(filename)
        out_path = self.music_cache_dir / safe_name
        try:
            out_path.write_bytes(data)
        except Exception as exc:
            messagebox.showerror("Feil", f"Kunne ikke skrive MP3: {exc}")
            return None
        return out_path

    def ensure_audio_backend(self):
        if self.audio_ready:
            return True
        try:
            import pygame
        except Exception:
            messagebox.showerror(
                "Feil",
                "For play/pause trengs pygame. Installer med: pip install pygame",
            )
            return False
        try:
            pygame.mixer.init()
        except Exception as exc:
            messagebox.showerror("Feil", f"Kunne ikke starte lyd: {exc}")
            return False
        self.audio_backend = pygame
        self.audio_ready = True
        return True

    def start_playback(self, path, filename):
        if not self.ensure_audio_backend():
            return False
        row = self.find_row_by_mp3(filename)
        self.current_duration = safe_int(row.get("MusikkSek")) if row else 0
        try:
            self.audio_backend.mixer.music.load(str(path))
            self.audio_backend.mixer.music.play()
        except Exception as exc:
            self.log(f"Kunne ikke spille av med intern avspiller: {exc}")
            try:
                try:
                    self.audio_backend.mixer.music.stop()
                except Exception:
                    pass
                os.startfile(str(path))
                self.external_playback = True
                self.is_paused = False
                self.play_pause_text.set("Spill")
                if row:
                    title = row.get("NavnFraIsonen") or row.get("PrintName") or ""
                    self.player_track_var.set(f"{title} - {filename}".strip(" -"))
                else:
                    self.player_track_var.set(filename)
                self.player_time_var.set("Ekstern avspiller")
                self.log(f"Spiller eksternt: {filename}")
                return True
            except Exception as exc2:
                messagebox.showerror("Feil", f"Kunne ikke spille av: {exc2}")
                return False
        self.current_track = filename
        self.is_paused = False
        self.external_playback = False
        self.play_pause_text.set("Pause")
        if row:
            title = row.get("NavnFraIsonen") or row.get("PrintName") or ""
            self.player_track_var.set(f"{title} - {filename}".strip(" -"))
        else:
            self.player_track_var.set(filename)
        self.log(f"Spiller: {filename}")
        return True

    def play_mp3_file(self, filename):
        path = self.get_cached_mp3_path(filename)
        if not path:
            return
        if self.ensure_audio_backend():
            self.start_playback(path, filename)
            return
        try:
            os.startfile(str(path))
            self.log(f"Spiller: {filename}")
        except Exception as exc:
            messagebox.showerror("Feil", f"Kunne ikke starte avspilling: {exc}")

    def get_selected_mp3_filename(self):
        selected = self.tree.selection()
        if not selected:
            return ""
        values = self.tree.item(selected[0], "values")
        columns = list(self.tree["columns"])
        try:
            col_index = columns.index("musikknavn")
        except ValueError:
            return ""
        if col_index >= len(values):
            return ""
        return values[col_index]

    def toggle_play_pause(self):
        if not self.ensure_audio_backend():
            return
        if self.external_playback:
            messagebox.showinfo(
                "Info",
                "Spiller i ekstern avspiller. Pause/fortsett er ikke tilgjengelig.",
            )
            return
        if self.audio_backend.mixer.music.get_busy():
            if self.is_paused:
                self.audio_backend.mixer.music.unpause()
                self.is_paused = False
                self.play_pause_text.set("Pause")
                self.log("Fortsetter avspilling.")
            else:
                self.audio_backend.mixer.music.pause()
                self.is_paused = True
                self.play_pause_text.set("Spill")
                self.log("Pause.")
            return
        filename = self.get_selected_mp3_filename()
        if not filename:
            messagebox.showinfo("Info", "Velg en rad med MP3-fil først.")
            return
        path = self.get_cached_mp3_path(filename)
        if not path:
            return
        self.start_playback(path, filename)

    def stop_playback(self):
        if not self.ensure_audio_backend():
            return
        self.audio_backend.mixer.music.stop()
        self.is_paused = False
        self.external_playback = False
        self.play_pause_text.set("Spill")
        self.log("Stoppet avspilling.")

    def find_row_by_mp3(self, filename):
        for row in self.rows:
            if row.get("MusikkFil") == filename:
                return row
        return None

    def update_player_ui(self):
        if self.external_playback:
            self.player_progress.config(maximum=1, value=0)
            self.root.after(500, self.update_player_ui)
            return
        elapsed = 0
        if self.audio_ready:
            pos_ms = self.audio_backend.mixer.music.get_pos()
            if pos_ms and pos_ms > 0:
                elapsed = pos_ms / 1000.0
        duration = self.current_duration or 0
        if duration > 0:
            self.player_progress.config(maximum=duration, value=min(elapsed, duration))
            elapsed_text = format_duration(elapsed)
            total_text = format_duration(duration)
            self.player_time_var.set(f"{elapsed_text} / {total_text}")
        else:
            self.player_progress.config(maximum=1, value=0)
            self.player_time_var.set("0:00 / 0:00")
        self.root.after(500, self.update_player_ui)
    
    def row_key(self, row):
        code = (row.get("ParticipantCode") or "").strip()
        if code:
            return f"code:{code}"
        given = (row.get("GivenName") or "").strip()
        family = (row.get("FamilyName") or "").strip()
        event = (row.get("Event") or "").strip()
        return f"name:{given}|{family}|{event}"

    def move_selected(self, delta):
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("Info", "Velg en linje i tabellen først.")
            return
        if len(self.rows) < 2:
            return
        idx = self.tree.index(selected[0])
        new_idx = idx + delta
        if new_idx < 0 or new_idx >= len(self.rows):
            return
        self.rows[idx], self.rows[new_idx] = self.rows[new_idx], self.rows[idx]
        self.refresh_table()
        new_item = self.tree.get_children()[new_idx]
        self.tree.selection_set(new_item)
        self.tree.see(new_item)

    def move_selected_up(self):
        self.move_selected(-1)

    def move_selected_down(self):
        self.move_selected(1)

    def shuffle_rows(self):
        if not self.rows:
            return
        random.shuffle(self.rows)
        self.refresh_table()

    def save_order(self):
        if not self.rows:
            return
        location = self.location_var.get().strip() or "sted"
        date_raw = self.start_date_var.get().strip()
        date_obj = parse_date_ddmmyy(date_raw)
        date_part = date_obj.strftime("%Y-%m-%d") if date_obj else date_raw.replace(".", "-")
        ts_display = datetime.now().strftime("%d.%m.%Y %H:%M")
        ts_filename = ts_display.replace(":", ".")
        base_name = f"{sanitize_filename(location)}_{sanitize_filename(date_part)}_{sanitize_filename(ts_filename)}_rekkefolge.json"
        initial_dir = str((self.zip_path.parent / "output") if self.zip_path else Path.cwd())
        path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("Rekkefølge", "*.json"), ("Alle filer", "*.*")],
            title="Lagre rekkefølge",
            initialfile=base_name,
            initialdir=initial_dir,
        )
        if not path:
            return
        data = {
            "version": 1,
            "created": datetime.now().isoformat(timespec="seconds"),
            "order": [self.row_key(r) for r in self.rows],
        }
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            self.log(f"Lagret rekkefølge: {path}")
        except Exception as exc:
            messagebox.showerror("Feil", f"Kunne ikke lagre rekkefølge: {exc}")

    def load_order(self):
        if not self.rows:
            return
        path = filedialog.askopenfilename(
            filetypes=[("Rekkefølge", "*.json"), ("Alle filer", "*.*")],
            title="Last rekkefølge",
        )
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as exc:
            messagebox.showerror("Feil", f"Kunne ikke lese rekkefølge: {exc}")
            return
        order = data.get("order") if isinstance(data, dict) else None
        if not order or not isinstance(order, list):
            messagebox.showerror("Feil", "Ugyldig rekkefølge-fil.")
            return
        buckets = {}
        for row in self.rows:
            key = self.row_key(row)
            buckets.setdefault(key, []).append(row)
        new_rows = []
        for key in order:
            bucket = buckets.get(key)
            if bucket:
                new_rows.append(bucket.pop(0))
        for bucket in buckets.values():
            new_rows.extend(bucket)
        self.rows = new_rows
        self.refresh_table()
        self.log(f"Lastet rekkefølge: {path}")

    def match_music_file(self, row, music_files, used_files):
        given = row.get("GivenName")
        family = row.get("FamilyName")
        given_tokens = tokenize_name(given)
        family_tokens = tokenize_name(family)
        if not family_tokens:
            return ""

        def family_match(fname):
            hay = normalize_text(fname)
            return all(token in hay for token in family_tokens) or family_tokens[0] in hay

        def given_match(fname):
            if not given_tokens:
                return False
            hay = normalize_text(fname)
            return given_tokens[0] in hay

        # Pass 1: require both family + given (if given exists), prefer unused.
        if given_tokens:
            for fname in music_files:
                if fname in used_files:
                    continue
                if family_match(fname) and given_match(fname):
                    used_files.add(fname)
                    return fname

        # Pass 2: family-only fallback, prefer unused.
        for fname in music_files:
            if fname in used_files:
                continue
            if family_match(fname):
                used_files.add(fname)
                return fname
        return ""

    def generate_files(self):
        if not self.rows or not self.zip_path:
            messagebox.showerror("Feil", "Ingen data lastet.")
            return

        out_dir = self.zip_path.parent / "output"
        out_dir.mkdir(parents=True, exist_ok=True)
        base_name = self.zip_path.stem

        if self.var_excel.get():
            generate_excel(self.rows, str(out_dir / f"{base_name}.xlsx"), self.log)
        if self.var_html.get():
            generate_html(
                self.rows, str(out_dir / f"{base_name}.html"), base_name, self.log
            )
        if self.var_pdf.get():
            generate_pdf(self.rows, str(out_dir / f"{base_name}.pdf"), base_name, self.log)

        self.log("Ferdig.")

    def generate_startliste(self):
        if not self.rows or not self.zip_path:
            messagebox.showerror("Feil", "Ingen data lastet.")
            return

        start_dt = parse_time_hhmm(self.start_time_var.get())
        if not start_dt:
            messagebox.showerror("Feil", "Ugyldig starttid. Bruk HH:MM.")
            return

        interval_seconds = parse_duration_mmss(self.interval_var.get())
        if not interval_seconds:
            messagebox.showerror("Feil", "Ugyldig intervall. Bruk M:SS eller H:MM:SS.")
            return
        warmup_seconds = parse_duration_mmss(self.warmup_var.get())
        if warmup_seconds is None:
            messagebox.showerror("Feil", "Ugyldig oppvarming. Bruk M:SS.")
            return

        try:
            group_size = int(self.group_size_var.get())
        except ValueError:
            messagebox.showerror("Feil", "Ugyldig gruppe-storrelse. Bruk et tall.")
            return
        if group_size <= 0:
            messagebox.showerror("Feil", "Gruppe-storrelse må være > 0.")
            return

        date_raw = self.start_date_var.get().strip()
        date_obj = parse_date_ddmmyy(date_raw)
        if not date_obj:
            messagebox.showerror("Feil", "Ugyldig dato. Bruk DD.MM.ÅÅ.")
            return
        date_text = format_date_long(date_obj)
        location = self.location_var.get().strip() or "iskanten"
        title = f"Oppvisningsstevne {location} {date_text}"

        filtered = [r for r in self.rows if is_registered(r.get("Påmelding", ""))]
        if not filtered:
            messagebox.showwarning("Info", "Fant ingen påmeldte i listen.")
            return
        for row in self.rows:
            row["StartTid"] = ""
        pause_after = None
        pause_seconds = None
        if self.pause_after_var.get().strip():
            try:
                pause_after = int(self.pause_after_var.get().strip())
            except ValueError:
                messagebox.showerror("Feil", "Ugyldig pause etter nr.")
                return
        if self.pause_duration_var.get().strip():
            pause_seconds = parse_duration_mmss(self.pause_duration_var.get().strip())
            if not pause_seconds:
                messagebox.showerror("Feil", "Ugyldig pause-varighet.")
                return
        pause_label = self.pause_label_var.get().strip() or "Vanningspause"

        entries = build_startliste(
            filtered,
            group_size,
            interval_seconds,
            start_dt,
            warmup_seconds=warmup_seconds,
            pause_after=pause_after,
            pause_seconds=pause_seconds,
            pause_label=pause_label,
        )
        filtered_index = 0
        for entry in entries:
            if entry.get("is_group"):
                continue
            if filtered_index >= len(filtered):
                break
            filtered[filtered_index]["StartTid"] = entry.get("start", "")
            filtered_index += 1
        self.refresh_table()
        out_dir = self.zip_path.parent / "output"
        out_dir.mkdir(parents=True, exist_ok=True)
        base_name = self.zip_path.stem
        generate_startliste_excel(
            entries, str(out_dir / f"Startliste_{base_name}.xlsx"), title, self.log
        )
        generate_startliste_pdf(
            entries, str(out_dir / f"Startliste_{base_name}.pdf"), title, self.log
        )
        if self.playlist_var.get():
            generate_vlc_playlist(filtered, out_dir, base_name, self.music_zip, self.log)
        self.log("Startliste ferdig.")

    def show_about(self):
        version = get_version()

        month_names = [
            "januar",
            "februar",
            "mars",
            "april",
            "mai",
            "juni",
            "juli",
            "august",
            "september",
            "oktober",
            "november",
            "desember",
        ]
        now = datetime.now()
        month_year = f"{month_names[now.month - 1]} {now.year}"

        about = tk.Toplevel(self.root)
        about.title("Om")
        about.resizable(False, False)

        frame = ttk.Frame(about, padding=16)
        frame.pack(fill="both", expand=True)

        logo_path = Path(__file__).resolve().parent / "Lil Logo.jpg"
        logo_label = None
        if logo_path.exists():
            try:
                from PIL import Image, ImageTk

                img = Image.open(logo_path)
                img.thumbnail((180, 180))
                photo = ImageTk.PhotoImage(img)
                logo_label = ttk.Label(frame, image=photo)
                logo_label.image = photo
                logo_label.pack(pady=(0, 8))
            except Exception:
                logo_label = None

        ttk.Label(
            frame,
            text="Loddefjord IL Kunstløp",
            font=("Segoe UI", 12, "bold"),
        ).pack()
        ttk.Label(
            frame,
            text=(
                "Programmet leser XML fra FSM-zip og lager "
                "Excel/HTML/PDF-utskrifter."
            ),
            wraplength=360,
            justify="center",
        ).pack(pady=(6, 2))
        ttk.Label(
            frame,
            text=f"Revisjon: {version} • {month_year}",
        ).pack(pady=(4, 0))


def main():
    root = tk.Tk()
    app = App(root)
    root.update_idletasks()
    screen_w = root.winfo_screenwidth()
    screen_h = root.winfo_screenheight()
    width = int(screen_w * 0.8)
    height = min(700, int(screen_h * 0.8))
    root.geometry(f"{width}x{height}")
    root.mainloop()


if __name__ == "__main__":
    main()
